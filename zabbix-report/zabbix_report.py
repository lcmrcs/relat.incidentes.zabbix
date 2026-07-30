"""
Gera relatórios de incidentes do Zabbix em Excel, HTML e PDF.

Este arquivo agora funciona como coordenador do fluxo. As regras de
classificação, acesso ao Zabbix, cálculo de indicadores e geração do PDF ficam
em módulos separados para facilitar leitura, manutenção e testes.
"""

import argparse
import base64
import logging
import os
import re
import unicodedata
from contextlib import nullcontext
from datetime import datetime, timedelta
from pathlib import Path

import pandas as pd
from classifiers import (
    SEVERITY_MAP,
    build_unit_catalog,
    classify_equipment,
    classify_incident_type,
    classify_unit_group,
)
from comparison import build_comparison_windows, build_executive_comparison
from data_integrity import validate_problem_records
from dotenv import load_dotenv
from jinja2 import Environment, FileSystemLoader, select_autoescape
from observability import ExecutionDiagnostics, write_optional_diagnostic
from openpyxl.chart import BarChart, DoughnutChart, Reference
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from pdf_report import technical_pdf_name, write_pdf_report, write_technical_pdf_report
from summary import build_report_summary, format_age
from time_utils import (
    datetime_to_unix,
    format_report_timestamp,
    now_display,
    parse_report_timestamp,
)
from zabbix_api import ZabbixClient

logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
LOGGER = logging.getLogger("zabbix-report")

# ==================================================
# CAMINHOS DO PROJETO
# ==================================================

BASE_DIR = Path(__file__).resolve().parent
ENV_FILE = BASE_DIR / ".env"
TEMPLATES_DIR = BASE_DIR / "templates"
ASSETS_DIR = BASE_DIR / "assets"
REPORTS_DIR = BASE_DIR / "reports"

EXCEL_COLUMNS = [
    ("date", "Data de abertura"),
    ("resolved_at", "Data de resolução"),
    ("unit_code", "Código da unidade"),
    ("unit", "Unidade"),
    ("host", "Host"),
    ("equipment", "Equipamento"),
    ("incident_type", "Tipo de incidente"),
    ("incident", "Incidente"),
    ("severity", "Severidade"),
    ("status", "Status"),
    ("duration_label", "Duração total"),
    ("open_age_label", "Idade do passivo aberto"),
    ("eventid", "Evento Zabbix"),
]


def parse_timestamp(value):
    """Converte timestamps Unix ou datas do relatório; valores inválidos viram None."""
    return parse_report_timestamp(value)


def format_timestamp(value):
    """Formata um timestamp válido sem interromper o relatório em dados corrompidos."""
    return format_report_timestamp(value)


def build_incident_key(unit_code, host, equipment, incident_type):
    """Cria a chave lógica determinística da mesma condição no mesmo ativo."""

    def normalize(value):
        text = unicodedata.normalize("NFKD", str(value or "").strip().casefold())
        text = text.encode("ascii", errors="ignore").decode("ascii")
        return re.sub(r"\s+", " ", text) or "n/a"

    return "|".join(normalize(value) for value in (unit_code, host, equipment, incident_type))


EXCEL_SEVERITY_COLORS = {
    "Desastre": "7F1D1D",
    "Alta": "EA580C",
    "Média": "D97706",
    "Atenção": "15803D",
    "Informação": "2563EB",
    "Não classificada": "64748B",
}


def slugify(value):
    """
    Converte textos livres em parte segura para nome de arquivo.

    Exemplo: "Terminal Facial" vira "terminal_facial".
    """

    normalized = str(value or "").strip().lower()
    normalized = unicodedata.normalize("NFKD", normalized)
    normalized = normalized.encode("ascii", errors="ignore").decode("ascii")
    normalized = re.sub(r"[^a-z0-9]+", "_", normalized)

    return normalized.strip("_") or "filtro"


def load_asset_data_uri(filename):
    """
    Carrega um arquivo de imagem em formato embutido para o HTML.

    O relatório final precisa continuar abrindo sozinho quando for enviado por
    e-mail ou copiado para outro computador. Por isso a imagem é transformada
    em base64, evitando dependência de um arquivo separado ao lado do HTML.
    """

    logo_path = ASSETS_DIR / filename
    if not logo_path.exists():
        return None

    mime_types = {
        ".png": "image/png",
        ".webp": "image/webp",
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
        ".svg": "image/svg+xml",
    }
    mime_type = mime_types.get(logo_path.suffix.lower(), "application/octet-stream")
    encoded_logo = base64.b64encode(logo_path.read_bytes()).decode("ascii")
    return f"data:{mime_type};base64,{encoded_logo}"


def load_logo_data_uri():
    """
    Carrega a logo principal do projeto em formato embutido para o HTML.
    """

    return load_asset_data_uri("logoTechPng.png")


def load_confea_logo_data_uri():
    """
    Carrega a logo da CONFEA para destacar o painel de VPN.
    """

    return load_asset_data_uri("logoConfea.png")


def load_zabbix_icon_data_uri():
    """
    Carrega o ícone oficial do Zabbix usado no painel resumido.
    """

    return load_asset_data_uri("zabbixLogoIcon.webp")


def load_zabbix_logo_data_uri():
    """
    Carrega a logo completa do Zabbix usada no modal de detalhes.
    """

    return load_asset_data_uri("zabbixLogoFull.png")


def cleanup_old_reports(current_base_name, keep_count=1):
    """
    Remove conjuntos antigos de relatórios gerados automaticamente.

    Cada execução cria um trio de arquivos com o mesmo nome base:
    .html, .pdf e .xlsx. Para evitar acumular arquivos antigos em
    zabbix-report/reports/, esta função mantém apenas os conjuntos mais
    recentes e apaga os anteriores.
    """

    try:
        keep_count = int(keep_count)
    except (TypeError, ValueError):
        keep_count = 1

    keep_count = max(1, keep_count)
    report_groups = {}

    for path in REPORTS_DIR.glob("report_*"):
        if path.suffix.lower() not in {".html", ".pdf", ".xlsx", ".json"}:
            continue

        # O anexo técnico pertence ao mesmo conjunto do PDF executivo e não
        # pode ser removido logo após uma execução com --pdf-detalhado.
        group_name = path.stem.removesuffix("_anexo_tecnico").removesuffix("_diagnostico")
        report_groups.setdefault(group_name, []).append(path)

    if current_base_name not in report_groups:
        report_groups[current_base_name] = []

    def group_mtime(item):
        _, paths = item
        if not paths:
            return 0
        return max(path.stat().st_mtime for path in paths if path.exists())

    ordered_groups = sorted(
        report_groups.items(),
        key=group_mtime,
        reverse=True,
    )
    keep_names = {current_base_name}

    for name, _ in ordered_groups:
        if len(keep_names) >= keep_count:
            break
        keep_names.add(name)
    removed = []

    for name, paths in report_groups.items():
        if name in keep_names:
            continue

        for path in paths:
            if path.exists():
                try:
                    path.unlink()
                    removed.append(path)
                except PermissionError:
                    print(
                        "AVISO: não foi possível remover relatório antigo "
                        f"porque o arquivo está em uso: {path}"
                    )

    return removed


# ==================================================
# ARGUMENTOS E PERÍODO
# ==================================================


def parse_period(value):
    """
    Converte textos como 24h, 2d e 7d em um timedelta.

    O valor "historico" retorna None para indicar que a consulta deve buscar
    desde o registro mais antigo disponível no Zabbix.
    """

    normalized = str(value).strip().lower()

    if normalized in ["historico", "histórico", "tudo", "todos", "all"]:
        return None, "histórico completo"

    if len(normalized) < 2:
        print("ERRO: use um período como 24h, 2d, 5d ou 7d.")
        raise SystemExit(1)

    amount = normalized[:-1]
    unit = normalized[-1]

    if not amount.isdigit() or int(amount) <= 0:
        print("ERRO: o valor do período precisa ser maior que zero.")
        raise SystemExit(1)

    amount = int(amount)

    if unit == "h":
        return timedelta(hours=amount), f"últimas {amount} hora(s)"

    if unit == "d":
        return timedelta(days=amount), f"últimos {amount} dia(s)"

    print("ERRO: unidade inválida. Use h para horas ou d para dias.")
    raise SystemExit(1)


def parse_args():
    """
    Lê as opções do terminal sem exigir alteração no código.

    --dias foi mantido por compatibilidade, mas --periodo é a opção mais
    flexível para 24h, 2d, 5d, 7d e historico.
    """

    parser = argparse.ArgumentParser(description="Gera relatórios de incidentes do Zabbix.")
    parser.add_argument(
        "--dias",
        type=int,
        default=None,
        help="Quantidade de dias que serão pesquisados. Mantido por compatibilidade.",
    )
    parser.add_argument(
        "--periodo",
        default="7d",
        help=("Intervalo pesquisado. Exemplos: 24h, 2d, 5d, 7d, 30d, historico. Padrão: 7d."),
    )
    parser.add_argument(
        "--desde",
        default=None,
        help="Data inicial no formato AAAA-MM-DD. Exemplo: --desde 2026-01-01.",
    )
    parser.add_argument(
        "--status",
        choices=["todos", "abertos", "resolvidos"],
        default="todos",
        help="Filtra eventos por situação. Use abertos para ignorar resolvidos.",
    )
    parser.add_argument(
        "--equipamento",
        default=None,
        help=(
            'Filtra o relatório por tipo de equipamento. Exemplo: --equipamento "Terminal Facial".'
        ),
    )
    parser.add_argument(
        "--unidade",
        default=None,
        help=("Filtra o relatório por código ou nome da unidade escolar. Exemplo: --unidade 1011."),
    )
    parser.add_argument(
        "--manter-relatorios",
        type=int,
        default=1,
        help=(
            "Quantidade de conjuntos antigos que devem permanecer na pasta "
            "reports. Padrão: 1, mantendo apenas o relatório mais recente."
        ),
    )
    parser.add_argument(
        "--pdf-detalhado",
        action="store_true",
        help=(
            "Gera, além do PDF executivo, um anexo técnico separado com todos "
            "os eventos do escopo."
        ),
    )
    parser.add_argument(
        "--diagnostico",
        action="store_true",
        help=(
            "Gera um JSON seguro com tempos, chamadas da API, tamanhos e " "gargalos da execução."
        ),
    )
    parser.add_argument(
        "--comparar",
        action="store_true",
        help="Compara o período selecionado com a janela anterior de mesma duração.",
    )

    args = parser.parse_args()

    if args.dias is not None and args.dias <= 0:
        print("ERRO: o argumento --dias precisa ser maior que zero.")
        raise SystemExit(1)

    if args.manter_relatorios <= 0:
        print("ERRO: o argumento --manter-relatorios precisa ser maior que zero.")
        raise SystemExit(1)

    return args


def resolve_period(args, today):
    """
    Calcula intervalo, rótulo e slug usados na API e no nome dos arquivos.
    """

    if args.desde:
        try:
            start_date = datetime.strptime(args.desde, "%Y-%m-%d")
        except ValueError:
            print("ERRO: use --desde no formato AAAA-MM-DD. Exemplo: 2026-01-01.")
            raise SystemExit(1) from None

        period_name = f"desde {start_date.strftime('%d/%m/%Y')}"
        period_slug = f"desde_{args.desde}"

    elif args.dias is not None:
        start_date = today - timedelta(days=args.dias)
        period_name = f"últimos {args.dias} dia(s)"
        period_slug = f"{args.dias}d"

    else:
        period_delta, period_name = parse_period(args.periodo)
        period_slug = str(args.periodo).strip().lower()
        start_date = today - period_delta if period_delta else None

    if args.status != "todos":
        period_name = f"{period_name} ({args.status})"
        period_slug = f"{period_slug}_{args.status}"

    return start_date, period_name, period_slug


def format_period_label(period_name, start_date, today):
    """
    Monta o texto exibido no HTML, PDF e terminal para o período consultado.
    """

    if start_date:
        return (
            f"{period_name}: {start_date.strftime('%d/%m/%Y %H:%M')} a "
            f"{today.strftime('%d/%m/%Y %H:%M')}"
        )

    return f"{period_name}: até {today.strftime('%d/%m/%Y %H:%M')}"


# ==================================================
# CONFIGURAÇÃO
# ==================================================


def load_config():
    """
    Carrega URL e token do arquivo .env local.

    O .env fica fora do Git porque contém credenciais. O script para cedo caso
    alguma variável obrigatória não exista.
    """

    load_dotenv(ENV_FILE)
    zabbix_url = os.getenv("ZABBIX_URL")
    zabbix_token = os.getenv("ZABBIX_TOKEN")

    if not zabbix_url or not zabbix_token:
        print("ERRO: Variáveis do .env não encontradas.")
        raise SystemExit(1)

    return zabbix_url, zabbix_token


def build_zabbix_web_url(zabbix_url):
    """
    Converte a URL da API em URL navegável do Zabbix.

    O .env guarda normalmente o endpoint JSON-RPC, como
    /api_jsonrpc.php. Para criar links clicáveis no relatório, removemos esse
    sufixo e usamos apenas a raiz web do Zabbix.
    """

    web_url = str(zabbix_url or "").strip()
    if web_url.endswith("/api_jsonrpc.php"):
        web_url = web_url[: -len("/api_jsonrpc.php")]

    return web_url.rstrip("/")


# ==================================================
# PROCESSAMENTO DOS INCIDENTES
# ==================================================


def build_incidents(
    problems,
    hosts_by_trigger,
    host_ids_by_trigger,
    host_details_by_id,
    resolved_at_by_event,
    unit_catalog,
    status_filter,
    generated_at,
):
    """
    Transforma problemas brutos do Zabbix em linhas prontas para relatório.

    Cada item retornado contém host, unidade, equipamento, incidente, severidade,
    status, datas e eventid. Esse formato único alimenta Excel, HTML e PDF.
    """

    incidents = []

    for item in problems:
        host = hosts_by_trigger.get(item.get("objectid"), "N/A")
        host_id = host_ids_by_trigger.get(item.get("objectid"))
        host_details = host_details_by_id.get(host_id, {})
        incident = item.get("name", "N/A")
        severity = SEVERITY_MAP.get(item.get("severity", "0"), "Desconhecida")
        timestamp = parse_timestamp(item.get("clock"))
        date = format_timestamp(timestamp)
        eventid = item.get("eventid")
        recovery_eventid = item.get("r_eventid")
        resolved_at = resolved_at_by_event.get(recovery_eventid, "")
        resolved_timestamp = parse_timestamp(resolved_at)
        status = "Resolvido" if resolved_timestamp is not None else "Aberto"

        generated_timestamp = datetime_to_unix(generated_at)
        duration_end = resolved_timestamp if status == "Resolvido" else generated_timestamp
        duration_seconds = max(0, duration_end - timestamp) if timestamp else 0
        open_age_seconds = duration_seconds if status == "Aberto" and timestamp else 0

        if status_filter == "abertos" and status != "Aberto":
            continue

        if status_filter == "resolvidos" and status != "Resolvido":
            continue

        equipment = classify_equipment(host)
        incident_type = classify_incident_type(incident)
        unit_code, unit = classify_unit_group(host, host_details, unit_catalog)

        incident_key = build_incident_key(unit_code, host, equipment, incident_type)

        incidents.append(
            {
                "host": host,
                "unit_code": unit_code,
                "unit": unit,
                "incident_key": incident_key,
                "equipment": equipment,
                "incident": incident,
                "incident_type": incident_type,
                "severity": severity,
                "status": status,
                "date": date,
                "timestamp": timestamp,
                # age_* permanece como alias de compatibilidade para filtros e templates.
                "age_seconds": open_age_seconds,
                "age_label": (
                    format_age(open_age_seconds)
                    if status == "Aberto"
                    else format_age(duration_seconds)
                ),
                "duration_seconds": duration_seconds,
                "duration_label": format_age(duration_seconds),
                "open_age_seconds": open_age_seconds,
                "open_age_label": format_age(open_age_seconds) if status == "Aberto" else "-",
                "resolved_timestamp": resolved_timestamp,
                "resolved_at": resolved_at,
                "eventid": eventid,
            }
        )

    return incidents


def split_special_groups(incidents):
    """
    Separa unidades escolares dos grupos especiais.

    Zabbix Server e CONFEA VPN são monitorados pelo Zabbix, mas não são unidade
    escolar. Separá-los mantém os indicadores escolares limpos.
    """

    main_incidents = [item for item in incidents if item["unit_code"] not in ["ZBX", "CONFEA"]]
    zabbix_incidents = [item for item in incidents if item["unit_code"] == "ZBX"]
    confea_incidents = [item for item in incidents if item["unit_code"] == "CONFEA"]

    return main_incidents, zabbix_incidents, confea_incidents


def filter_by_equipment(incidents, equipment_name):
    """
    Mantém apenas incidentes do equipamento informado no argumento.

    A comparação ignora maiúsculas/minúsculas e espaços extras para facilitar o
    uso no terminal.
    """

    if not equipment_name:
        return incidents

    target = str(equipment_name).strip().lower()

    return [item for item in incidents if str(item.get("equipment", "")).strip().lower() == target]


def filter_by_unit(incidents, unit_filter):
    """
    Mantém apenas incidentes da unidade escolar informada.

    O filtro aceita tanto o código numérico quanto parte do nome da unidade.
    Isso deixa o uso mais simples na tela inicial e no terminal.
    """

    if not unit_filter:
        return incidents

    target = str(unit_filter).strip().lower()

    return [
        item
        for item in incidents
        if (
            target in str(item.get("unit_code", "")).strip().lower()
            or target in str(item.get("unit", "")).strip().lower()
        )
    ]


# ==================================================
# EXPORTAÇÃO
# ==================================================


def incidents_to_excel_frame(incidents):
    """
    Converte incidentes em DataFrame com nomes amigáveis para o Excel.

    A lista de colunas é fixa para manter a planilha previsível mesmo quando
    alguma execução não retorna incidentes.
    """

    rows = []

    for item in incidents:
        rows.append({label: item.get(key, "") for key, label in EXCEL_COLUMNS})

    return pd.DataFrame(rows, columns=[label for _, label in EXCEL_COLUMNS])


def counter_to_excel_frame(items):
    """
    Transforma rankings do resumo em uma tabela simples para o Excel.
    """

    return pd.DataFrame(
        [
            {
                "Nome": item["name"],
                "Total": item["total"],
                "Percentual": item["percent"],
            }
            for item in items
        ],
        columns=["Nome", "Total", "Percentual"],
    )


def build_excel_summary_rows(summary, generated, period_label):
    """
    Monta os blocos textuais da aba Resumo Executivo.
    """

    age = summary["age"]

    return [
        ("Relatório Executivo de Incidentes Zabbix", ""),
        ("Gerado em", generated),
        ("Período analisado", period_label),
        ("Produzido por", "Network Operations Center"),
        ("", ""),
        ("Eventos totais", summary["event_total"]),
        ("Incidentes únicos", summary["unique_total"]),
        ("Eventos repetidos", summary["repeated_events"]),
        ("Eventos abertos", summary["open"]),
        ("Eventos resolvidos", summary["resolved"]),
        ("Incidentes abertos", summary["unique_open"]),
        ("Mais antigo aberto", age["oldest_label"]),
        ("Média de idade", age["average_label"]),
        ("Acima de 7 dias", age["over_7d"]),
        ("Índice médio de prioridade", summary["priority"]["average_score"]),
        ("Prioridade crítica", summary["priority"]["critical"]),
        ("Prioridade alta", summary["priority"]["high"]),
        ("Hosts reincidentes", summary["recurrence"]["affected_hosts"]),
        ("", ""),
        ("Alta", summary["high"]),
        ("Média", summary["medium"]),
        ("Atenção", summary["attention"]),
        ("Informação", summary["information"]),
        ("Desastre", summary["critical"]),
    ]


def build_excel_intelligence_frames(summary):
    """
    Monta tabelas executivas de comparativo, reincidência e prioridade.
    """

    comparison = pd.DataFrame(
        [
            {
                "Faixa": item["label"],
                "Total": item["total"],
                "Percentual": item["percent"],
                "Alta criticidade": item["high"],
            }
            for item in summary["period_comparison"]["ranges"]
        ],
        columns=["Faixa", "Total", "Percentual", "Alta criticidade"],
    )
    recurrence = pd.DataFrame(
        [
            {
                "Host": item["host"],
                "Unidade": item["unit"],
                "Equipamento": item["equipment"],
                "Tipo de incidente": item["incident_type"],
                "Ocorrências": item["total"],
                "Índice": item["score"],
            }
            for item in summary["recurrence"]["top"]
        ],
        columns=[
            "Host",
            "Unidade",
            "Equipamento",
            "Tipo de incidente",
            "Ocorrências",
            "Índice",
        ],
    )
    priority = pd.DataFrame(
        [
            {
                "Índice": item["score"],
                "Prioridade": item["label"],
                "Host": item["host"],
                "Unidade": item["unit"],
                "Equipamento": item["equipment"],
                "Tipo de incidente": item["incident_type"],
                "Severidade": item["severity"],
                "Idade do passivo aberto": item["open_age_label"],
                "Evento": item["eventid"],
            }
            for item in summary["priority"]["top"]
        ],
        columns=[
            "Índice",
            "Prioridade",
            "Host",
            "Unidade",
            "Equipamento",
            "Tipo de incidente",
            "Severidade",
            "Idade do passivo aberto",
            "Evento",
        ],
    )

    return [
        ("Distribuição temporal", comparison),
        ("Padrões recorrentes", recurrence),
        ("Prioridades operacionais", priority),
        (
            "Duração histórica dos resolvidos",
            pd.DataFrame(
                [
                    {
                        "Faixa": "Total resolvidos",
                        "Total": summary["resolved_duration"]["total"],
                        "Percentual": "",
                    },
                    {
                        "Faixa": "Duração média",
                        "Total": summary["resolved_duration"]["average_label"],
                        "Percentual": "",
                    },
                    {
                        "Faixa": "Duração mediana",
                        "Total": summary["resolved_duration"]["median_label"],
                        "Percentual": "",
                    },
                    {
                        "Faixa": "Maior duração",
                        "Total": summary["resolved_duration"]["maximum_label"],
                        "Percentual": "",
                    },
                ]
                + [
                    {
                        "Faixa": item["label"],
                        "Total": item["total"],
                        "Percentual": item["percent"],
                    }
                    for item in summary["resolved_duration"]["ranges"]
                ],
                columns=["Faixa", "Total", "Percentual"],
            ),
        ),
    ]


def build_integrity_frame(integrity):
    """Cria a visão tabular canônica de integridade usada pelo Excel."""

    rows = [
        {
            "Categoria": "Registros recebidos",
            "Quantidade": integrity["received"],
            "Tratamento": "Validação",
            "Impacto": "Base da coleta",
        },
        {
            "Categoria": "Registros processados",
            "Quantidade": integrity["processed"],
            "Tratamento": "Mantidos",
            "Impacto": "Usados no relatório",
        },
        {
            "Categoria": "Registros ajustados",
            "Quantidade": integrity["adjusted"],
            "Tratamento": "Fallback seguro",
            "Impacto": "Mantidos com normalização",
        },
        {
            "Categoria": "Registros descartados",
            "Quantidade": integrity["discarded"],
            "Tratamento": "Descartados",
            "Impacto": "Relatório pode ficar incompleto",
        },
    ]
    rows.extend(
        {
            "Categoria": item["category"],
            "Quantidade": item["quantity"],
            "Tratamento": item["treatment"],
            "Impacto": item["impact"],
        }
        for item in integrity["issues"]
    )
    return pd.DataFrame(rows, columns=["Categoria", "Quantidade", "Tratamento", "Impacto"])


def build_excel_criticality_frame(summary):
    """
    Cria a visão executiva de criticidade por unidade escolar.

    Essa aba leva para o Excel o mesmo raciocínio do mapa do HTML: pontuação,
    faixa operacional, volume, severidade, reincidência e equipamentos afetados.
    """

    rows = []

    for item in summary["unit_criticality"]["top"]:
        equipment_mix = ", ".join(
            f"{equipment['name']}: {equipment['total']}" for equipment in item["equipment_mix"]
        )
        severity_mix = ", ".join(
            f"{severity['name']}: {severity['total']}" for severity in item["severity_mix"]
        )
        factors = item["factors"]

        rows.append(
            {
                "Score": item["score"],
                "Faixa operacional": item["level"],
                "Código": item["code"],
                "Unidade": item["name"],
                "Incidentes": item["total"],
                "Equipamentos afetados": item["affected_equipment_count"],
                "Principal equipamento": (
                    f"{item['top_equipment']} ({item['top_equipment_total']})"
                ),
                "Severidade predominante": (
                    f"{item['top_severity']} ({item['top_severity_total']})"
                ),
                "Mais antigo": item["oldest_label"],
                "Média offline": item["age_label"],
                "Reincidência": item["recurrence"],
                "Composição dos equipamentos": equipment_mix,
                "Composição da severidade": severity_mix,
                "Fator volume": f"{factors['volume']}%",
                "Fator severidade": f"{factors['severity']}%",
                "Fator tempo offline": f"{factors['age']}%",
                "Fator reincidência": f"{factors['recurrence']}%",
                "Fator equipamento": f"{factors['equipment']}%",
            }
        )

    return pd.DataFrame(
        rows,
        columns=[
            "Score",
            "Faixa operacional",
            "Código",
            "Unidade",
            "Incidentes",
            "Equipamentos afetados",
            "Principal equipamento",
            "Severidade predominante",
            "Mais antigo",
            "Média offline",
            "Reincidência",
            "Composição dos equipamentos",
            "Composição da severidade",
            "Fator volume",
            "Fator severidade",
            "Fator tempo offline",
            "Fator reincidência",
            "Fator equipamento",
        ],
    )


def excel_level_colors(level):
    """
    Define cores da faixa operacional usada na aba de criticidade.
    """

    colors = {
        "Intervenção imediata": ("991B1B", "FFFFFF"),
        "Prioridade alta": ("EA580C", "FFFFFF"),
        "Acompanhamento ativo": ("D97706", "FFFFFF"),
        "Monitoramento normal": ("0F766E", "FFFFFF"),
    }
    return colors.get(str(level), ("64748B", "FFFFFF"))


def style_excel_workbook(writer, diagnostics=None):
    """
    Aplica acabamento visual, filtros e congelamento em todas as abas.

    A formatação fica no final para que os dados sejam exportados primeiro pelo
    pandas e depois refinados com openpyxl.
    """

    workbook = writer.book
    workbook.properties.title = "Relatório Executivo de Incidentes Zabbix"
    workbook.properties.subject = "Monitoramento operacional via Zabbix"
    workbook.properties.creator = "Network Operations Center"

    header_fill = PatternFill("solid", fgColor="073B43")
    accent_fill = PatternFill("solid", fgColor="E8FAF8")
    dark_fill = PatternFill("solid", fgColor="062A30")
    soft_fill = PatternFill("solid", fgColor="F6FBFB")
    section_fill = PatternFill("solid", fgColor="0F766E")
    warning_fill = PatternFill("solid", fgColor="FFF7ED")
    danger_fill = PatternFill("solid", fgColor="FEE2E2")
    border_color = "BFD8DC"
    thin_border = Border(
        left=Side(style="thin", color=border_color),
        right=Side(style="thin", color=border_color),
        top=Side(style="thin", color=border_color),
        bottom=Side(style="thin", color=border_color),
    )
    measure = diagnostics.measure if diagnostics else lambda _name: nullcontext()

    def register_style(name, font=None, fill=None, alignment=None):
        if name in workbook.named_styles:
            named_style = workbook._named_styles[name]
        else:
            named_style = NamedStyle(
                name=name,
                font=font or Font(),
                fill=fill or PatternFill(),
                border=thin_border,
                alignment=alignment or Alignment(vertical="top", wrap_text=True),
            )
            workbook.add_named_style(named_style)
        return named_style._style

    header_style = register_style(
        "NocHeader",
        font=Font(color="FFFFFF", bold=True),
        fill=header_fill,
        alignment=Alignment(horizontal="center", vertical="center"),
    )
    body_style = register_style("NocBody")
    body_even_style = register_style("NocBodyEven", fill=soft_fill)
    severity_styles = {
        severity: register_style(
            f"NocSeverity{index}",
            font=Font(color="FFFFFF", bold=True),
            fill=PatternFill("solid", fgColor=color),
            alignment=Alignment(horizontal="center", vertical="top", wrap_text=True),
        )
        for index, (severity, color) in enumerate(EXCEL_SEVERITY_COLORS.items())
    }
    status_styles = {
        "Aberto": register_style(
            "NocStatusOpen",
            font=Font(color="B91C1C", bold=True),
            fill=PatternFill("solid", fgColor="FEE2E2"),
        ),
        "Resolvido": register_style(
            "NocStatusResolved",
            font=Font(color="166534", bold=True),
            fill=PatternFill("solid", fgColor="DCFCE7"),
        ),
    }
    level_styles = {}
    for index, level in enumerate(
        (
            "Intervenção imediata",
            "Prioridade alta",
            "Acompanhamento ativo",
            "Monitoramento normal",
        )
    ):
        fill_color, font_color = excel_level_colors(level)
        level_styles[level] = register_style(
            f"NocLevel{index}",
            font=Font(color=font_color, bold=True),
            fill=PatternFill("solid", fgColor=fill_color),
            alignment=Alignment(horizontal="center", vertical="center"),
        )
    score_styles = {
        False: register_style(
            "NocScore",
            font=Font(color="073B43", bold=True, size=13),
            alignment=Alignment(horizontal="center", vertical="center"),
        ),
        True: register_style(
            "NocScoreEven",
            font=Font(color="073B43", bold=True, size=13),
            fill=soft_fill,
            alignment=Alignment(horizontal="center", vertical="center"),
        ),
    }

    with measure("excel_base_styles"):
        for worksheet in workbook.worksheets:
            worksheet.sheet_view.showGridLines = False
            worksheet.freeze_panes = "A2"
            worksheet.sheet_view.zoomScale = 90
            for cell in worksheet[1]:
                cell._style = header_style
            for row in worksheet.iter_rows(min_row=2):
                row_style = body_even_style if row[0].row % 2 == 0 else body_style
                for cell in row:
                    cell._style = row_style
            worksheet.row_dimensions[1].height = 24

    with measure("excel_column_widths"):
        for worksheet in workbook.worksheets:
            for column_cells in worksheet.columns:
                column_letter = get_column_letter(column_cells[0].column)
                max_length = max(len(str(cell.value or "")) for cell in column_cells[:80])
                worksheet.column_dimensions[column_letter].width = min(
                    max(max_length + 3, 13),
                    42,
                )

    data_sheets = [
        (sheet_index, worksheet)
        for sheet_index, worksheet in enumerate(workbook.worksheets)
        if worksheet.title not in {"Resumo Executivo", "Rankings", "Inteligência", "Comparativo"}
        and worksheet.max_row > 1
        and worksheet.max_column > 1
    ]
    with measure("excel_tables"):
        for sheet_index, worksheet in data_sheets:
            worksheet.auto_filter.ref = worksheet.dimensions
            table = Table(displayName=f"Tabela{sheet_index + 1}", ref=worksheet.dimensions)
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            worksheet.add_table(table)

    with measure("excel_conditional_formatting"):
        for _, worksheet in data_sheets:
            headers = [cell.value for cell in worksheet[1]]
            severity_index = headers.index("Severidade") + 1 if "Severidade" in headers else None
            status_index = headers.index("Status") + 1 if "Status" in headers else None
            level_index = (
                headers.index("Faixa operacional") + 1 if "Faixa operacional" in headers else None
            )
            score_index = headers.index("Score") + 1 if "Score" in headers else None

            for row in worksheet.iter_rows(min_row=2):
                if severity_index:
                    severity_cell = row[severity_index - 1]
                    severity_cell._style = severity_styles.get(
                        str(severity_cell.value),
                        body_even_style if severity_cell.row % 2 == 0 else body_style,
                    )

                if status_index:
                    status_cell = row[status_index - 1]
                    if status_cell.value in status_styles:
                        status_cell._style = status_styles[status_cell.value]

                if level_index:
                    level_cell = row[level_index - 1]
                    if level_cell.value in level_styles:
                        level_cell._style = level_styles[level_cell.value]

                if score_index:
                    score_cell = row[score_index - 1]
                    score_cell._style = score_styles[score_cell.row % 2 == 0]

    summary_sheet = workbook["Resumo Executivo"]
    summary_sheet.freeze_panes = None
    summary_sheet.sheet_view.zoomScale = 95
    summary_sheet.column_dimensions["A"].width = 32
    summary_sheet.column_dimensions["B"].width = 44
    summary_sheet["A1"].fill = dark_fill
    summary_sheet["B1"].fill = dark_fill
    summary_sheet["A1"].font = Font(color="FFFFFF", bold=True, size=16)
    summary_sheet["B1"].font = Font(color="FFFFFF", bold=True, size=16)

    for row_number in range(2, 20):
        summary_sheet[f"A{row_number}"].font = Font(color="455A64", bold=True)
        summary_sheet[f"B{row_number}"].font = Font(color="073B43", bold=True)
        summary_sheet[f"A{row_number}"].fill = accent_fill
        summary_sheet[f"B{row_number}"].fill = accent_fill

    for row_number in (6, 7, 8, 9, 10, 11, 12, 13):
        summary_sheet[f"B{row_number}"].font = Font(
            color="087F8C",
            bold=True,
            size=13,
        )

    for row_number in (15, 16, 17, 18, 19):
        summary_sheet[f"B{row_number}"].font = Font(
            color="EA580C" if row_number in (15, 16) else "087F8C",
            bold=True,
            size=12,
        )

    summary_sheet.sheet_properties.tabColor = "087F8C"

    for worksheet in workbook.worksheets:
        if worksheet.title == "Unidades":
            worksheet.sheet_properties.tabColor = "087F8C"
        elif worksheet.title == "Servidor Zabbix":
            worksheet.sheet_properties.tabColor = "DC2626"
        elif worksheet.title == "CONFEA VPN":
            worksheet.sheet_properties.tabColor = "7C3AED"
        elif worksheet.title == "Todos":
            worksheet.sheet_properties.tabColor = "0F766E"
        elif worksheet.title == "Criticidade":
            worksheet.sheet_properties.tabColor = "EA580C"

    if "Rankings" in workbook.sheetnames:
        rankings_sheet = workbook["Rankings"]
        rankings_sheet.sheet_properties.tabColor = "0E7490"
        rankings_sheet.freeze_panes = None

        for row in rankings_sheet.iter_rows():
            first_cell = row[0]
            if first_cell.value and all(cell.value in (None, "") for cell in row[1:]):
                for cell in row[:3]:
                    cell.fill = section_fill
                    cell.font = Font(color="FFFFFF", bold=True)
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                rankings_sheet.row_dimensions[first_cell.row].height = 22

    if "Inteligência" in workbook.sheetnames:
        intelligence_sheet = workbook["Inteligência"]
        intelligence_sheet.sheet_properties.tabColor = "12343B"
        intelligence_sheet.freeze_panes = None

        for row in intelligence_sheet.iter_rows():
            first_cell = row[0]
            if first_cell.value and all(cell.value in (None, "") for cell in row[1:]):
                for cell in row[: max(1, intelligence_sheet.max_column)]:
                    cell.fill = dark_fill
                    cell.font = Font(color="FFFFFF", bold=True)
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                intelligence_sheet.row_dimensions[first_cell.row].height = 22

    if "Criticidade" in workbook.sheetnames:
        criticality_sheet = workbook["Criticidade"]
        criticality_sheet.freeze_panes = "D2"
        criticality_sheet.sheet_view.zoomScale = 85
        criticality_sheet.column_dimensions["A"].width = 12
        criticality_sheet.column_dimensions["B"].width = 24
        criticality_sheet.column_dimensions["D"].width = 42
        criticality_sheet.column_dimensions["G"].width = 28
        criticality_sheet.column_dimensions["H"].width = 28
        criticality_sheet.column_dimensions["L"].width = 34
        criticality_sheet.column_dimensions["M"].width = 34

        for row in criticality_sheet.iter_rows(min_row=2):
            score = row[0].value or 0
            if score >= 82:
                row[0].fill = danger_fill
                row[0].font = Font(color="991B1B", bold=True, size=13)
            elif score >= 64:
                row[0].fill = warning_fill
                row[0].font = Font(color="EA580C", bold=True, size=13)
            else:
                row[0].fill = accent_fill
                row[0].font = Font(color="0F766E", bold=True, size=13)
            row[0].alignment = Alignment(horizontal="center", vertical="center")


def add_excel_charts(writer):
    """
    Cria gráficos simples na aba de resumo a partir da aba Rankings.
    """

    workbook = writer.book

    if "Rankings" not in workbook.sheetnames:
        return

    summary_sheet = workbook["Resumo Executivo"]
    rankings_sheet = workbook["Rankings"]

    if rankings_sheet.max_row < 3:
        return

    chart = DoughnutChart()
    chart.title = "Severidade"
    labels = Reference(rankings_sheet, min_col=1, min_row=2, max_row=6)
    data = Reference(rankings_sheet, min_col=2, min_row=1, max_row=6)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)
    chart.holeSize = 58
    chart.height = 7
    chart.width = 9
    summary_sheet.add_chart(chart, "D2")

    equipment_start = 10
    equipment_end = min(rankings_sheet.max_row, equipment_start + 7)

    if equipment_end > equipment_start:
        bar_chart = BarChart()
        bar_chart.title = "Equipamentos mais afetados"
        bar_chart.y_axis.title = "Incidentes"
        bar_chart.x_axis.title = "Equipamento"
        labels = Reference(
            rankings_sheet,
            min_col=1,
            min_row=equipment_start + 1,
            max_row=equipment_end,
        )
        data = Reference(
            rankings_sheet,
            min_col=2,
            min_row=equipment_start,
            max_row=equipment_end,
        )
        bar_chart.add_data(data, titles_from_data=True)
        bar_chart.set_categories(labels)
        bar_chart.height = 7
        bar_chart.width = 12
        summary_sheet.add_chart(bar_chart, "D18")


def export_excel(
    path,
    all_incidents,
    main_incidents,
    zabbix_incidents,
    confea_incidents,
    summary,
    generated,
    period_label,
    integrity_summary=None,
    diagnostics=None,
    comparison=None,
):
    """
    Gera a planilha Excel com abas separadas por finalidade.

    A aba Unidades é a visão escolar. As abas Servidor Zabbix e CONFEA VPN
    isolam infraestrutura especial. A aba Todos preserva a visão completa.
    """

    measure = diagnostics.measure if diagnostics else lambda _name: nullcontext()
    integrity_summary = integrity_summary or {
        "received": len(all_incidents),
        "processed": len(all_incidents),
        "adjusted": 0,
        "discarded": 0,
        "issues": [],
    }

    with measure("excel_dataframes"):
        summary_frame = pd.DataFrame(
            build_excel_summary_rows(summary, generated, period_label),
            columns=["Indicador", "Valor"],
        )
        rankings_frames = [
            ("Severidade", counter_to_excel_frame(summary["severity"])),
            ("Equipamentos", counter_to_excel_frame(summary["top_equipment"])),
            ("Tipos de incidente", counter_to_excel_frame(summary["top_incident_types"])),
            ("Unidades", counter_to_excel_frame(summary["top_units"])),
            ("Hosts", counter_to_excel_frame(summary["top_hosts"])),
        ]
        intelligence_frames = build_excel_intelligence_frames(summary)
        criticality_frame = build_excel_criticality_frame(summary)
        integrity_frame = build_integrity_frame(integrity_summary)
        main_frame = incidents_to_excel_frame(main_incidents)
        zabbix_frame = incidents_to_excel_frame(zabbix_incidents) if zabbix_incidents else None
        confea_frame = incidents_to_excel_frame(confea_incidents) if confea_incidents else None
        all_frame = incidents_to_excel_frame(all_incidents)
        comparison_frame = None
        comparison_changes = []
        if comparison:
            comparison_frame = pd.DataFrame(
                [
                    {
                        "Indicador": item["label"],
                        "Período atual": item["current_label"],
                        "Período anterior": item["previous_label"],
                        "Diferença": item["difference_label"],
                        "Variação": item["percent_label"],
                        "Direção": item["direction"],
                        "Interpretação": item["interpretation"],
                    }
                    for item in comparison["metrics"]
                ]
            )
            for title, key in (
                ("Mudanças por unidade", "unit_changes"),
                ("Mudanças por equipamento", "equipment_changes"),
            ):
                comparison_changes.append(
                    (
                        title,
                        pd.DataFrame(
                            comparison[key],
                            columns=["name", "current", "previous", "difference"],
                        ).rename(
                            columns={
                                "name": "Nome",
                                "current": "Atual",
                                "previous": "Anterior",
                                "difference": "Diferença",
                            }
                        ),
                    )
                )

    writer = pd.ExcelWriter(path, engine="openpyxl")
    try:
        with measure("excel_sheet_writes"):
            summary_frame.to_excel(
                writer,
                sheet_name="Resumo Executivo",
                index=False,
            )
            start_row = 0

            for title, frame in rankings_frames:
                pd.DataFrame(
                    [[title, "", ""]],
                    columns=["Nome", "Total", "Percentual"],
                ).to_excel(
                    writer,
                    sheet_name="Rankings",
                    index=False,
                    header=start_row == 0,
                    startrow=start_row,
                )
                frame.to_excel(
                    writer,
                    sheet_name="Rankings",
                    index=False,
                    header=False,
                    startrow=start_row + 1,
                )
                start_row += len(frame) + 4

            start_row = 0
            for title, frame in intelligence_frames:
                pd.DataFrame([[title]], columns=["Indicador"]).to_excel(
                    writer,
                    sheet_name="Inteligência",
                    index=False,
                    header=False,
                    startrow=start_row,
                )
                frame.to_excel(
                    writer,
                    sheet_name="Inteligência",
                    index=False,
                    startrow=start_row + 1,
                )
                start_row += len(frame) + 4

            criticality_frame.to_excel(
                writer,
                sheet_name="Criticidade",
                index=False,
            )
            integrity_frame.to_excel(
                writer,
                sheet_name="Integridade dos Dados",
                index=False,
            )
            if comparison_frame is not None:
                pd.DataFrame(
                    [
                        ["Período atual", comparison["current_label"]],
                        ["Período anterior", comparison["previous_label"]],
                        ["Qualidade", comparison["quality_label"]],
                        ["Integridade", comparison.get("integrity_note", "")],
                        ["Escopo", comparison["note"]],
                    ],
                    columns=["Comparativo executivo", "Valor"],
                ).to_excel(writer, sheet_name="Comparativo", index=False)
                comparison_frame.to_excel(
                    writer,
                    sheet_name="Comparativo",
                    index=False,
                    startrow=7,
                )
                comparison_row = 9 + len(comparison_frame)
                for title, frame in comparison_changes:
                    pd.DataFrame([[title]], columns=["Indicador"]).to_excel(
                        writer,
                        sheet_name="Comparativo",
                        index=False,
                        header=False,
                        startrow=comparison_row,
                    )
                    frame.to_excel(
                        writer,
                        sheet_name="Comparativo",
                        index=False,
                        startrow=comparison_row + 1,
                    )
                    comparison_row += len(frame) + 4
            main_frame.to_excel(
                writer,
                sheet_name="Unidades",
                index=False,
            )
            if zabbix_frame is not None:
                zabbix_frame.to_excel(
                    writer,
                    sheet_name="Servidor Zabbix",
                    index=False,
                )
            if confea_frame is not None:
                confea_frame.to_excel(
                    writer,
                    sheet_name="CONFEA VPN",
                    index=False,
                )
            all_frame.to_excel(
                writer,
                sheet_name="Todos",
                index=False,
            )

        style_excel_workbook(writer, diagnostics=diagnostics)
        with measure("excel_charts"):
            add_excel_charts(writer)
        with measure("excel_save"):
            writer.close()
    except Exception:
        writer.book.close()
        raise


def render_html(
    path,
    generated,
    period_label,
    main_incidents,
    summary,
    zabbix_incidents,
    zabbix_summary,
    confea_incidents,
    confea_summary,
    zabbix_web_url,
    integrity_summary=None,
    comparison=None,
):
    """
    Renderiza o template HTML com os dados já processados.

    O Jinja2 separa apresentação dos dados: o Python prepara informações e o
    template decide como mostrar cards, filtros, tabelas e janelas.
    """

    env = Environment(
        loader=FileSystemLoader(TEMPLATES_DIR),
        autoescape=select_autoescape(["html", "xml"]),
    )
    template = env.get_template("report_template.html")
    html_output = template.render(
        generated=generated,
        period=period_label,
        total=len(main_incidents),
        incidents=main_incidents,
        summary=summary,
        zabbix_incidents=zabbix_incidents,
        zabbix_summary=zabbix_summary,
        confea_incidents=confea_incidents,
        confea_summary=confea_summary,
        logo_data_uri=load_logo_data_uri(),
        zabbix_icon_data_uri=load_zabbix_icon_data_uri(),
        zabbix_logo_data_uri=load_zabbix_logo_data_uri(),
        confea_logo_data_uri=load_confea_logo_data_uri(),
        zabbix_web_url=zabbix_web_url,
        integrity=integrity_summary
        or {
            "level": "valid",
            "label": "Dados validados",
            "warning_count": 0,
            "received": len(main_incidents),
            "processed": len(main_incidents),
            "adjusted": 0,
            "discarded": 0,
            "issues": [],
        },
        incident_payload=build_html_incident_payload(main_incidents),
        comparison=comparison,
    )

    with open(path, "w", encoding="utf-8") as file:
        file.write(html_output)


def build_html_incident_payload(incidents):
    """Compacta os incidentes em uma única fonte de dados para o navegador."""

    fields = (
        "date",
        "resolved_at",
        "status",
        "unit_code",
        "unit",
        "host",
        "equipment",
        "incident",
        "incident_type",
        "severity",
        "timestamp",
        "age_seconds",
        "age_label",
        "duration_seconds",
        "duration_label",
        "open_age_seconds",
        "open_age_label",
        "eventid",
    )
    return [[item.get(field, "") for field in fields] for item in incidents]


# ==================================================
# EXECUÇÃO PRINCIPAL
# ==================================================


def execute_measured_export(diagnostics, stage, kind, path, callback, with_pages=False):
    """Executa uma exportação, registra duração, arquivo e falha segura."""

    try:
        with diagnostics.measure(stage):
            result = callback()
    except Exception as error:
        diagnostics.record_failed_file(kind, path, error)
        LOGGER.error(
            "Falha na etapa %s: %s",
            diagnostics.stages[stage]["label"],
            type(error).__name__,
        )
        raise

    diagnostics.record_file(
        kind,
        path,
        pages=result if with_pages else None,
    )
    LOGGER.info("%s concluída", diagnostics.stages[stage]["label"])
    return result


def requested_period_label(args):
    """Resume o período solicitado sem incluir nomes de filtros operacionais."""

    if args.desde:
        period = f"desde {args.desde}"
    elif args.dias is not None:
        period = f"{args.dias}d"
    else:
        period = str(args.periodo)
    comparison_label = " · comparação ativa" if getattr(args, "comparar", False) else ""
    return f"{period} · status {args.status}{comparison_label}"


def main():
    """
    Coordena o relatório inteiro, do terminal até os arquivos finais.
    """

    args = parse_args()
    diagnostics = ExecutionDiagnostics(requested_period_label(args))

    with diagnostics.measure("configuration"):
        zabbix_url, zabbix_token = load_config()
        zabbix_web_url = build_zabbix_web_url(zabbix_url)
        today = now_display()
        start_date, period_name, period_slug = resolve_period(args, today)
        period_label = format_period_label(period_name, start_date, today)
        time_from = datetime_to_unix(start_date) if start_date else None
        time_till = datetime_to_unix(today)
        comparison_windows = None
        if getattr(args, "comparar", False):
            if start_date is None:
                print(
                    "ERRO: --comparar exige um período finito. "
                    "Use --periodo 24h, --periodo 7d ou --desde AAAA-MM-DD."
                )
                raise SystemExit(1)
            comparison_windows = build_comparison_windows(start_date, today)

    REPORTS_DIR.mkdir(exist_ok=True)
    base_name = f"report_{today.strftime('%Y-%m-%d')}_{period_slug}"
    generated = today.strftime("%d/%m/%Y %H:%M")
    excel_name = REPORTS_DIR / f"{base_name}.xlsx"
    html_name = REPORTS_DIR / f"{base_name}.html"
    pdf_name = REPORTS_DIR / f"{base_name}.pdf"
    technical_pdf = technical_pdf_name(pdf_name)
    diagnostic_name = REPORTS_DIR / f"{base_name}_diagnostico.json"

    try:
        client = ZabbixClient(zabbix_url, zabbix_token, diagnostics=diagnostics)

        LOGGER.info("Iniciando coleta de eventos do Zabbix")
        print("Conectando ao Zabbix...")
        with diagnostics.measure("api_collection"):
            with diagnostics.measure("problem_search"):
                collection_status = "todos" if comparison_windows else args.status
                problems = client.get_problems(collection_status, time_from, time_till)
            LOGGER.info("Coleta concluída: %d registro(s) recebido(s)", len(problems))
            with diagnostics.measure("recovery_search"):
                resolved_at_by_event = client.get_recovery_dates(problems)
            with diagnostics.measure("trigger_host_search"):
                (
                    hosts_by_trigger,
                    host_ids_by_trigger,
                    host_details_by_id,
                ) = client.get_trigger_hosts(problems)
            with diagnostics.measure("unit_catalog"):
                all_host_details_by_id = client.get_all_hosts_with_tags()
                unit_catalog = build_unit_catalog(all_host_details_by_id)
            previous_problems = []
            previous_resolved_at_by_event = {}
            previous_hosts_by_trigger = {}
            previous_host_ids_by_trigger = {}
            previous_host_details_by_id = {}
            if comparison_windows:
                LOGGER.info("Iniciando coleta do período anterior para comparação")
                with diagnostics.measure("comparison_previous_collection"):
                    with diagnostics.measure("comparison_previous_problem_search"):
                        previous_problems = client.get_problems(
                            "todos",
                            comparison_windows["previous"]["time_from"],
                            comparison_windows["previous"]["time_till"],
                        )
                    with diagnostics.measure("comparison_previous_recovery_search"):
                        previous_resolved_at_by_event = client.get_recovery_dates(previous_problems)
                    with diagnostics.measure("comparison_previous_trigger_host_search"):
                        (
                            previous_hosts_by_trigger,
                            previous_host_ids_by_trigger,
                            previous_host_details_by_id,
                        ) = client.get_trigger_hosts(previous_problems)
                LOGGER.info(
                    "Coleta comparativa concluída: %d registro(s) no período anterior",
                    len(previous_problems),
                )

        print("Processando incidentes...")
        with diagnostics.measure("validation"):
            problems, integrity_summary = validate_problem_records(
                problems,
                hosts_by_trigger,
                host_ids_by_trigger,
                host_details_by_id,
                resolved_at_by_event,
                unit_catalog,
                today,
            )
        diagnostics.set_record_counts(integrity_summary)
        LOGGER.info(
            "Validação concluída: %d processado(s), %d ajustado(s), %d descartado(s)",
            integrity_summary["processed"],
            integrity_summary["adjusted"],
            integrity_summary["discarded"],
        )
        if integrity_summary["warning_count"]:
            LOGGER.warning(
                "Foram identificados %d aviso(s) de integridade",
                integrity_summary["warning_count"],
            )
        previous_integrity_summary = None
        if comparison_windows:
            with diagnostics.measure("comparison_previous_validation"):
                previous_problems, previous_integrity_summary = validate_problem_records(
                    previous_problems,
                    previous_hosts_by_trigger,
                    previous_host_ids_by_trigger,
                    previous_host_details_by_id,
                    previous_resolved_at_by_event,
                    unit_catalog,
                    comparison_windows["previous"]["end"],
                )

        with diagnostics.measure("incident_build"):
            incidents = build_incidents(
                problems,
                hosts_by_trigger,
                host_ids_by_trigger,
                host_details_by_id,
                resolved_at_by_event,
                unit_catalog,
                args.status,
                today,
            )
            main_incidents, zabbix_incidents, confea_incidents = split_special_groups(incidents)
            equipment_filter = str(args.equipamento).strip() if args.equipamento else ""
            unit_filter = str(args.unidade).strip() if args.unidade else ""

            if unit_filter:
                main_incidents = filter_by_unit(main_incidents, unit_filter)
                zabbix_incidents = []
                confea_incidents = []
                period_label = f"{period_label} | Unidade: {unit_filter}"
                period_slug = f"{period_slug}_unidade_{slugify(unit_filter)}"

            if equipment_filter:
                main_incidents = filter_by_equipment(main_incidents, equipment_filter)
                zabbix_incidents = []
                confea_incidents = []
                period_label = f"{period_label} | Equipamento: {equipment_filter}"
                period_slug = f"{period_slug}_{slugify(equipment_filter)}"

            scoped_incidents = main_incidents + zabbix_incidents + confea_incidents
            comparison = None
            if comparison_windows:
                comparison_current_incidents = build_incidents(
                    problems,
                    hosts_by_trigger,
                    host_ids_by_trigger,
                    host_details_by_id,
                    resolved_at_by_event,
                    unit_catalog,
                    "todos",
                    today,
                )
                comparison_current_main, _, _ = split_special_groups(comparison_current_incidents)
                previous_incidents = build_incidents(
                    previous_problems,
                    previous_hosts_by_trigger,
                    previous_host_ids_by_trigger,
                    previous_host_details_by_id,
                    previous_resolved_at_by_event,
                    unit_catalog,
                    "todos",
                    comparison_windows["previous"]["end"],
                )
                previous_main, _, _ = split_special_groups(previous_incidents)
                if unit_filter:
                    comparison_current_main = filter_by_unit(comparison_current_main, unit_filter)
                    previous_main = filter_by_unit(previous_main, unit_filter)
                if equipment_filter:
                    comparison_current_main = filter_by_equipment(
                        comparison_current_main, equipment_filter
                    )
                    previous_main = filter_by_equipment(previous_main, equipment_filter)
                comparison = build_executive_comparison(
                    comparison_current_main + previous_main,
                    previous_main + comparison_current_main,
                    comparison_windows,
                    integrity_summary,
                    previous_integrity_summary,
                )

        # Filtros alteram o slug e, portanto, os nomes finais dos arquivos.
        base_name = f"report_{today.strftime('%Y-%m-%d')}_{period_slug}"
        excel_name = REPORTS_DIR / f"{base_name}.xlsx"
        html_name = REPORTS_DIR / f"{base_name}.html"
        pdf_name = REPORTS_DIR / f"{base_name}.pdf"
        technical_pdf = technical_pdf_name(pdf_name)
        diagnostic_name = REPORTS_DIR / f"{base_name}_diagnostico.json"
        planned_files = {
            "excel": excel_name,
            "html": html_name,
            "pdf": pdf_name,
        }
        if args.pdf_detalhado:
            planned_files["technical_pdf"] = technical_pdf
        for kind, path in planned_files.items():
            diagnostics.record_file(kind, path, completed=False)

        with diagnostics.measure("summaries"):
            summary = build_report_summary(main_incidents)
            zabbix_summary = build_report_summary(zabbix_incidents)
            confea_summary = build_report_summary(confea_incidents)
        diagnostics.set_event_groups(
            summary["event_total"],
            zabbix_summary["event_total"],
            confea_summary["event_total"],
        )

        execute_measured_export(
            diagnostics,
            "excel_export",
            "excel",
            excel_name,
            lambda: export_excel(
                excel_name,
                scoped_incidents,
                main_incidents,
                zabbix_incidents,
                confea_incidents,
                summary,
                generated,
                period_label,
                integrity_summary,
                diagnostics,
                comparison,
            ),
        )
        print(f"Excel gerado: {excel_name}")

        execute_measured_export(
            diagnostics,
            "html_export",
            "html",
            html_name,
            lambda: render_html(
                html_name,
                generated,
                period_label,
                main_incidents,
                summary,
                zabbix_incidents,
                zabbix_summary,
                confea_incidents,
                confea_summary,
                zabbix_web_url,
                integrity_summary,
                comparison,
            ),
        )
        print(f"HTML gerado: {html_name}")

        execute_measured_export(
            diagnostics,
            "pdf_export",
            "pdf",
            pdf_name,
            lambda: write_pdf_report(
                pdf_name,
                main_incidents,
                generated,
                summary,
                period_label,
                integrity_summary,
                {
                    "Servidor Zabbix": zabbix_summary,
                    "CONFEA VPN": confea_summary,
                },
                comparison,
            ),
            with_pages=True,
        )
        print(f"PDF gerado: {pdf_name}")

        if args.pdf_detalhado:
            execute_measured_export(
                diagnostics,
                "technical_pdf_export",
                "technical_pdf",
                technical_pdf,
                lambda: write_technical_pdf_report(
                    technical_pdf,
                    scoped_incidents,
                    generated,
                ),
                with_pages=True,
            )
            print(f"Anexo técnico PDF gerado: {technical_pdf}")

        with diagnostics.measure("report_cleanup"):
            removed_reports = cleanup_old_reports(
                base_name,
                keep_count=args.manter_relatorios,
            )
        if removed_reports:
            print(f"Relatórios antigos removidos: {len(removed_reports)} arquivo(s).")

    except (Exception, SystemExit):
        diagnostics.finalize()
        if write_optional_diagnostic(args.diagnostico, diagnostics, diagnostic_name):
            LOGGER.info("Diagnóstico parcial salvo: %s", diagnostic_name.name)
        diagnostics.print_terminal_summary()
        raise

    diagnostics.finalize()
    if write_optional_diagnostic(args.diagnostico, diagnostics, diagnostic_name):
        print(f"Diagnóstico gerado: {diagnostic_name}")

    print("\nRELATÓRIOS GERADOS COM SUCESSO")
    diagnostics.print_terminal_summary()
    LOGGER.info("Geração do relatório concluída em %.2fs", diagnostics.total_seconds)


if __name__ == "__main__":
    main()
