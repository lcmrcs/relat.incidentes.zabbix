import json
import re
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

PROJECT_DIR = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_DIR))

from observability import ExecutionDiagnostics  # noqa: E402
from pdf_report import write_pdf_report  # noqa: E402
from summary import build_report_summary  # noqa: E402
from zabbix_report import export_excel, render_html  # noqa: E402


def sample_incident():
    return {
        "host": "1040_23-TERM_FACIAL - PORTARIA",
        "unit_code": "1040",
        "unit": "1040-CETI de Catu",
        "incident_key": "1040|host|Terminal Facial|High ICMP ping loss|Alta",
        "equipment": "Terminal Facial",
        "incident": "High ICMP ping loss - 10.0.0.1",
        "incident_type": "High ICMP ping loss",
        "severity": "Alta",
        "status": "Aberto",
        "date": "26/06/2026 14:16",
        "timestamp": 1782480000,
        "age_seconds": 3600,
        "age_label": "1h 0min",
        "duration_seconds": 3600,
        "duration_label": "1h 0min",
        "open_age_seconds": 3600,
        "open_age_label": "1h 0min",
        "resolved_at": "",
        "eventid": "123456",
    }


class ExportTests(unittest.TestCase):
    def test_export_excel_creates_executive_sheets(self):
        incidents = [sample_incident()]
        summary = build_report_summary(incidents)

        with tempfile.TemporaryDirectory() as temp_dir:
            output = Path(temp_dir) / "relatorio.xlsx"
            export_excel(
                output,
                incidents,
                incidents,
                [],
                [],
                summary,
                "26/06/2026 14:16",
                "histórico completo (abertos): até 26/06/2026 14:16",
            )

            workbook = load_workbook(output)

        self.assertEqual(
            workbook.sheetnames,
            [
                "Resumo Executivo",
                "Rankings",
                "Inteligência",
                "Criticidade",
                "Integridade dos Dados",
                "Unidades",
                "Todos",
            ],
        )
        self.assertEqual(workbook["Resumo Executivo"]["A3"].value, "Gerado em")
        self.assertEqual(workbook["Inteligência"]["A1"].value, "Distribuição temporal")
        self.assertEqual(workbook["Criticidade"]["A1"].value, "Score")
        self.assertEqual(workbook["Criticidade"]["B1"].value, "Faixa operacional")
        self.assertEqual(workbook["Integridade dos Dados"]["A1"].value, "Categoria")
        self.assertEqual(workbook["Unidades"]["A1"].value, "Data de abertura")
        self.assertEqual(workbook["Unidades"]["B1"].value, "Data de resolução")
        self.assertEqual(workbook["Unidades"]["K1"].value, "Duração total")
        self.assertEqual(workbook["Unidades"]["L1"].value, "Idade do passivo aberto")
        self.assertEqual(workbook["Unidades"]["M1"].value, "Evento Zabbix")

    def test_export_excel_reports_internal_optimization_stages(self):
        incidents = [sample_incident()]
        summary = build_report_summary(incidents)
        diagnostics = ExecutionDiagnostics("período fictício")

        with tempfile.TemporaryDirectory() as temp_dir:
            output = Path(temp_dir) / "relatorio.xlsx"
            export_excel(
                output,
                incidents,
                incidents,
                [],
                [],
                summary,
                "26/06/2026 14:16",
                "período fictício",
                diagnostics=diagnostics,
            )

        measured_stages = {item["name"] for item in diagnostics.as_dict()["stages"]}
        self.assertTrue(
            {
                "excel_dataframes",
                "excel_sheet_writes",
                "excel_base_styles",
                "excel_column_widths",
                "excel_tables",
                "excel_conditional_formatting",
                "excel_charts",
                "excel_save",
            }.issubset(measured_stages)
        )

    def test_write_pdf_report_creates_pdf_file(self):
        incidents = [sample_incident()]
        summary = build_report_summary(incidents)

        with tempfile.TemporaryDirectory() as temp_dir:
            output = Path(temp_dir) / "relatorio.pdf"
            write_pdf_report(
                output,
                incidents,
                "26/06/2026 14:16",
                summary,
                "histórico completo (abertos): até 26/06/2026 14:16",
            )
            pdf_bytes = output.read_bytes()

        self.assertTrue(pdf_bytes.startswith(b"%PDF-1.4"))
        self.assertIn(b"%%EOF", pdf_bytes)
        self.assertIn(b"1h 0min", pdf_bytes)

    def test_html_excel_and_pdf_consume_the_same_operational_fields(self):
        resolved = {
            **sample_incident(),
            "incident_key": "1040|host|terminal facial|high icmp ping loss resolved",
            "eventid": "123457",
            "status": "Resolvido",
            "resolved_at": "26/06/2026 16:16",
            "duration_seconds": 7200,
            "duration_label": "2h 0min",
            "open_age_seconds": 0,
            "open_age_label": "-",
            "age_seconds": 0,
            "age_label": "2h 0min",
        }
        incidents = [sample_incident(), resolved]
        summary = build_report_summary(incidents)
        integrity = {
            "received": 3,
            "processed": 2,
            "adjusted": 1,
            "discarded": 1,
            "duplicates": 1,
            "invalid_timestamps": 0,
            "inconsistent_recoveries": 0,
            "unidentified_hosts": 0,
            "unidentified_units": 0,
            "unidentified_equipment": 0,
            "unknown_severities": 0,
            "warning_count": 1,
            "level": "incomplete",
            "label": "Relatório possivelmente incompleto",
            "issues": [
                {
                    "key": "duplicate_eventid",
                    "category": "Evento duplicado",
                    "quantity": 1,
                    "treatment": "Duplicata descartada",
                    "impact": "Evita contagem repetida.",
                }
            ],
        }

        with tempfile.TemporaryDirectory() as temp_dir:
            temp_dir = Path(temp_dir)
            html_path = temp_dir / "relatorio.html"
            excel_path = temp_dir / "relatorio.xlsx"
            pdf_path = temp_dir / "relatorio.pdf"

            render_html(
                html_path,
                "26/06/2026 14:16",
                "período fictício",
                incidents,
                summary,
                [],
                build_report_summary([]),
                [],
                build_report_summary([]),
                "https://zabbix.example.test",
                integrity,
            )
            export_excel(
                excel_path,
                incidents,
                incidents,
                [],
                [],
                summary,
                "26/06/2026 14:16",
                "período fictício",
                integrity,
            )
            write_pdf_report(
                pdf_path,
                incidents,
                "26/06/2026 14:16",
                summary,
                "período fictício",
                integrity,
            )

            html = html_path.read_text(encoding="utf-8")
            workbook = load_workbook(excel_path, data_only=True)
            pdf = pdf_path.read_bytes()

        payload_match = re.search(
            r'<script id="incident-data" type="application/json">(.*?)</script>',
            html,
            re.DOTALL,
        )
        self.assertIsNotNone(payload_match)
        payload = json.loads(payload_match.group(1))

        self.assertEqual(payload[0][14], "1h 0min")
        self.assertIn("Relatório possivelmente incompleto", html)
        self.assertNotIn("Duração dos incidentes resolvidos", html)
        self.assertNotIn("resolved-duration-track", html)
        self.assertEqual(payload[0][16], "1h 0min")
        self.assertEqual(payload[1][14], "2h 0min")
        self.assertEqual(payload[1][16], "-")
        self.assertIn("data-page-status", html)
        self.assertEqual(workbook["Unidades"]["K2"].value, "1h 0min")
        self.assertEqual(workbook["Unidades"]["L2"].value, "1h 0min")
        self.assertEqual(workbook["Unidades"]["K3"].value, "2h 0min")
        self.assertEqual(workbook["Unidades"]["L3"].value, "-")
        self.assertEqual(workbook["Integridade dos Dados"]["B5"].value, 1)
        self.assertIn(b"1h 0min", pdf)
        self.assertIn(b"2h 0min", pdf)
        self.assertIn(b"possivelmente incompleto", pdf)
        self.assertIn(b"mediana", pdf)


if __name__ == "__main__":
    unittest.main()
